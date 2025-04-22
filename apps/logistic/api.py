from fastapi import APIRouter, HTTPException
from tortoise.transactions import in_transaction
from apps.logistic.models import AreaCode, PostPrice
from pathlib import Path
from typing import List, Dict
import openpyxl  # 使用openpyxl代替pandas

router = APIRouter()


async def get_existing_records() -> List[Dict]:
    """获取所有已存在的记录的三字段组合"""
    records = await AreaCode.all().values('country_code', 'ship_code',
                                          'post_code')
    return [{
        'country_code': r['country_code'],
        'ship_code': r['ship_code'],
        'post_code': r['post_code']
    } for r in records]


@router.get("/import_area_code/", summary="导入物流分区代码")
async def import_area_code():
    excel_path = Path("media/load/area_code.xlsx")

    if not excel_path.exists():
        raise HTTPException(status_code=404, detail="Excel文件不存在")

    try:
        # 使用openpyxl读取Excel文件
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active

        # 获取标题行
        headers = [cell.value for cell in sheet[1]]
        required_columns = [
            'country_code', 'name', 'ship_code', 'post_code', 'area', 'service'
        ]

        # 检查必要列是否存在
        missing_columns = [
            col for col in required_columns if col not in headers
        ]
        if missing_columns:
            raise HTTPException(
                status_code=400,
                detail=f"Excel缺少必要列: {', '.join(missing_columns)}")

        # 准备批量创建的数据
        to_create = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            to_create.append(
                AreaCode(country_code=row[headers.index('country_code')],
                         name=row[headers.index('name')],
                         ship_code=row[headers.index('ship_code')],
                         post_code=row[headers.index('post_code')],
                         area=row[headers.index('area')],
                         is_service=False
                         if str(row[headers.index('service')]).lower()
                         == "out of network" else True))

        # 先删除所有数据再批量创建
        async with in_transaction():
            await AreaCode.all().delete()
            if to_create:
                await AreaCode.bulk_create(to_create)

        return {
            "status": "success",
            "created": len(to_create),
            "deleted": "all"  # 表示已删除所有旧数据
        }

    except Exception as e:
        raise HTTPException(status_code=400, detail=f"导入失败: {str(e)}")


@router.get("/import_post_price/", summary="导入物流价格数据")
async def import_post_price():
    excel_path = Path("media/load/post_price.xlsx")

    if not excel_path.exists():
        raise HTTPException(status_code=404, detail="Excel文件不存在")

    try:
        # 使用openpyxl读取Excel文件
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active

        # 获取标题行
        headers = [cell.value for cell in sheet[1]]
        required_columns = [
            'country_code', 'carrier_name', 'carrier_code', 'min_weight',
            'max_weight', 'area', 'basic_price', 'calc_price', 'volume_ratio',
            'is_elec'
        ]

        # 检查必要列是否存在
        missing_columns = [
            col for col in required_columns if col not in headers
        ]
        if missing_columns:
            raise HTTPException(
                status_code=400,
                detail=f"Excel缺少必要列: {', '.join(missing_columns)}")

        # 准备批量创建的数据
        to_create = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            to_create.append(
                PostPrice(country_code=row[headers.index('country_code')],
                          carrier_name=row[headers.index('carrier_name')],
                          carrier_code=row[headers.index('carrier_code')],
                          min_weight=row[headers.index('min_weight')],
                          max_weight=row[headers.index('max_weight')],
                          area=row[headers.index('area')],
                          basic_price=row[headers.index('basic_price')],
                          calc_price=row[headers.index('calc_price')],
                          volume_ratio=row[headers.index('volume_ratio')],
                          is_elec=bool(row[headers.index('is_elec')])))

        # 先删除所有数据再批量创建
        async with in_transaction():
            await PostPrice.all().delete()
            if to_create:
                await PostPrice.bulk_create(to_create)

        return {
            "status": "success",
            "created": len(to_create),
            "deleted": "all"  # 表示已删除所有旧数据
        }

    except Exception as e:
        raise HTTPException(status_code=400, detail=f"导入失败: {str(e)}")
